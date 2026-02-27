"""Balance upload, summary, details, contracts, progress, export, and delete routes."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd
from fastapi import APIRouter, File, HTTPException, Response, UploadFile
from fastapi.responses import StreamingResponse

import app.state as state
from app.schemas import (
    BalanceContract,
    BalanceContractsResponse,
    BalanceDetailsResponse,
    BalanceUploadResponse,
)
from app.session import (
    _assert_session_exists,
    _calc_params_path,
    _motor_positions_path,
    _positions_path,
    _results_path,
    _session_dir,
    _summary_path,
)
from app.parsers.balance_parser import (
    _load_or_rebuild_positions_df,
    _load_or_rebuild_summary,
    _parse_and_store_balance,
    _parse_zip_balance,
)
from app.parsers._persistence import _invalidate_positions_cache
from app.parsers.transforms import _to_float, _to_text
from engine.banks import default_bank
from app.services.balance_query import (
    _aggregate_groups_df,
    _aggregate_totals_df,
    _apply_filters_df,
    _build_cross_filtered_facets_df,
)

router = APIRouter()


@router.delete("/api/sessions/{session_id}/balance")
def delete_balance(session_id: str) -> dict[str, str]:
    _assert_session_exists(session_id)
    _invalidate_positions_cache(session_id)
    sdir = _session_dir(session_id)
    deleted: list[str] = []
    positions_parquet = _positions_path(session_id)
    positions_json_legacy = positions_parquet.with_suffix(".json")
    motor_parquet = _motor_positions_path(session_id)
    motor_json_legacy = motor_parquet.with_suffix(".json")
    contracts_json_orphan = sdir / "balance_contracts.json"
    for p in [
        _summary_path(session_id), positions_parquet, positions_json_legacy,
        motor_parquet, motor_json_legacy, contracts_json_orphan,
    ]:
        if p.exists():
            p.unlink()
            deleted.append(p.name)
    for p in sdir.iterdir():
        if p.is_file() and p.name.startswith("balance__"):
            p.unlink()
            deleted.append(p.name)
    for p in [_results_path(session_id), _calc_params_path(session_id)]:
        if p.exists():
            p.unlink()
            deleted.append(p.name)
    return {"status": "ok", "deleted": ", ".join(deleted) if deleted else "nothing to delete"}


@router.post("/api/sessions/{session_id}/balance", response_model=BalanceUploadResponse)
async def upload_balance(session_id: str, file: UploadFile = File(...)) -> BalanceUploadResponse:
    _assert_session_exists(session_id)
    _invalidate_positions_cache(session_id)

    raw_filename = file.filename or "balance.xlsx"
    if not raw_filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xls files are supported")

    safe_filename = Path(raw_filename).name
    storage_name = f"balance__{safe_filename}"

    sdir = _session_dir(session_id)
    xlsx_path = sdir / storage_name
    content = await file.read()
    xlsx_path.write_bytes(content)

    return _parse_and_store_balance(session_id, filename=safe_filename, xlsx_path=xlsx_path)


@router.post("/api/sessions/{session_id}/balance/zip", response_model=BalanceUploadResponse)
async def upload_balance_zip(
    session_id: str,
    file: UploadFile = File(...),
    bank_id: str = default_bank(),
) -> BalanceUploadResponse:
    import asyncio

    from engine.banks import resolve_bank as resolve_adapter

    _assert_session_exists(session_id)
    _invalidate_positions_cache(session_id)

    try:
        adapter = resolve_adapter(bank_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    raw_filename = file.filename or "balance.zip"
    if not raw_filename.lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="Only .zip files are supported for this endpoint")

    safe_filename = Path(raw_filename).name
    sdir = _session_dir(session_id)
    zip_path = sdir / f"balance__{safe_filename}"
    content = await file.read()
    zip_path.write_bytes(content)

    # Run in thread so the event loop stays responsive for progress polling.
    # Tree building + persistence happen inside the thread to eliminate race
    # conditions (files are written before the HTTP response is sent).
    response = await asyncio.to_thread(
        _parse_zip_balance, session_id, zip_path, safe_filename, adapter=adapter,
    )
    state._upload_progress.pop(session_id, None)
    return response


@router.get("/api/sessions/{session_id}/upload-progress")
def get_upload_progress(session_id: str, response: Response):
    response.headers["Cache-Control"] = "no-store"
    progress = state._upload_progress.get(session_id)
    if progress is None:
        return {"phase": "idle", "step": 0, "total": 0, "pct": 0, "phase_label": ""}
    step = progress.get("step", 0)
    total = progress.get("total", 0)
    phase = progress.get("phase", "parsing")

    # (phase_name, label, pct_start, pct_span)
    _PHASE_CONFIG = [
        ("parsing",        "Parsing positions\u2026",      5,  60),
        ("persisting",     "Saving motor data\u2026",      65,  7),
        ("canonicalizing", "Building aggregates\u2026",    72, 10),
        ("building_tree",  "Building summary tree\u2026",  82,  8),
        ("saving",         "Saving positions\u2026",       90,  8),
    ]

    pct = 5
    phase_label = "Processing\u2026"
    for name, label, pct_start, pct_span in _PHASE_CONFIG:
        if phase == name:
            phase_label = label
            progress_frac = (step / total) if total > 0 else 0.5
            pct = pct_start + round(progress_frac * pct_span)
            break

    return {"phase": phase, "step": step, "total": total, "pct": pct, "phase_label": phase_label}


@router.get("/api/sessions/{session_id}/balance/summary", response_model=BalanceUploadResponse)
def get_balance_summary(session_id: str) -> BalanceUploadResponse:
    _assert_session_exists(session_id)
    return _load_or_rebuild_summary(session_id)


@router.get("/api/sessions/{session_id}/balance/details", response_model=BalanceDetailsResponse)
def get_balance_details(
    session_id: str,
    categoria_ui: str | None = None,
    subcategoria_ui: str | None = None,
    subcategory_id: str | None = None,
    currency: str | None = None,
    rate_type: str | None = None,
    counterparty: str | None = None,
    segment: str | None = None,
    strategic_segment: str | None = None,
    maturity: str | None = None,
    remuneration: str | None = None,
    book_value: str | None = None,
    group_by: str | None = None,
) -> BalanceDetailsResponse:
    _assert_session_exists(session_id)

    df = _load_or_rebuild_positions_df(session_id)

    context_df = _apply_filters_df(
        df,
        categoria_ui=categoria_ui,
        subcategoria_ui=subcategoria_ui,
        subcategory_id=subcategory_id,
    )

    filtered_df = _apply_filters_df(
        context_df,
        currency=currency,
        rate_type=rate_type,
        counterparty=counterparty,
        segment=segment,
        strategic_segment=strategic_segment,
        maturity=maturity,
        remuneration=remuneration,
        book_value=book_value,
    )

    group_by_cols = [c.strip() for c in group_by.split(",")] if group_by else None
    groups = _aggregate_groups_df(filtered_df, group_by=group_by_cols)
    totals = _aggregate_totals_df(filtered_df)
    # Cross-filtered facets: each dimension's counts reflect all OTHER active
    # filters but NOT its own.  This way selecting "<1Y" maturity still shows
    # all maturity options (with counts narrowed by currency/segment/etc.),
    # and users can freely multi-select within any dimension.
    facets = _build_cross_filtered_facets_df(
        context_df,
        currency=currency,
        rate_type=rate_type,
        segment=segment,
        strategic_segment=strategic_segment,
        maturity=maturity,
        remuneration=remuneration,
        book_value=book_value,
    )

    pretty_subcategory = subcategoria_ui
    if pretty_subcategory is None and subcategory_id:
        match = context_df.loc[
            context_df["subcategory_id"].fillna("").str.lower() == subcategory_id.lower()
        ]
        if not match.empty:
            val = match["subcategoria_ui"].iloc[0]
            pretty_subcategory = str(val) if pd.notna(val) else subcategory_id

    return BalanceDetailsResponse(
        session_id=session_id,
        categoria_ui=categoria_ui,
        subcategoria_ui=pretty_subcategory,
        groups=groups,
        totals=totals,
        facets=facets,
    )


@router.get("/api/sessions/{session_id}/balance/contracts", response_model=BalanceContractsResponse)
def get_balance_contracts(
    session_id: str,
    query: str | None = None,
    categoria_ui: str | None = None,
    subcategoria_ui: str | None = None,
    subcategory_id: str | None = None,
    group: str | None = None,
    currency: str | None = None,
    rate_type: str | None = None,
    counterparty: str | None = None,
    segment: str | None = None,
    strategic_segment: str | None = None,
    maturity: str | None = None,
    remuneration: str | None = None,
    book_value: str | None = None,
    page: int = 1,
    page_size: int = 100,
    offset: int | None = None,
    limit: int | None = None,
) -> BalanceContractsResponse:
    _assert_session_exists(session_id)

    if page < 1:
        raise HTTPException(status_code=400, detail="page must be >= 1")

    if offset is not None or limit is not None:
        effective_offset = max(offset or 0, 0)
        effective_limit = limit or 200
        if effective_limit <= 0 or effective_limit > 2000:
            raise HTTPException(status_code=400, detail="limit must be between 1 and 2000")
        effective_page = (effective_offset // effective_limit) + 1
        effective_page_size = effective_limit
    else:
        effective_page = page
        effective_page_size = page_size

    if effective_page_size <= 0 or effective_page_size > 2000:
        raise HTTPException(status_code=400, detail="page_size must be between 1 and 2000")

    df = _load_or_rebuild_positions_df(session_id)

    filtered = _apply_filters_df(
        df,
        categoria_ui=categoria_ui,
        subcategoria_ui=subcategoria_ui,
        subcategory_id=subcategory_id,
        group=group,
        currency=currency,
        rate_type=rate_type,
        counterparty=counterparty,
        segment=segment,
        strategic_segment=strategic_segment,
        maturity=maturity,
        remuneration=remuneration,
        book_value=book_value,
        query_text=query,
    )

    total = len(filtered)
    start = (effective_page - 1) * effective_page_size
    end = start + effective_page_size

    # Only convert the paginated slice to dicts (100-200 rows, not 1.5M)
    sliced = filtered.iloc[start:end]
    sliced_records = sliced.where(sliced.notna(), other=None).to_dict("records")
    contracts = [
        BalanceContract(
            contract_id=str(row.get("contract_id") or ""),
            sheet=_to_text(row.get("sheet")),
            category=str(row.get("side") or ""),
            categoria_ui=_to_text(row.get("categoria_ui")),
            subcategory=str(row.get("subcategory_id") or "unknown"),
            subcategoria_ui=_to_text(row.get("subcategoria_ui")),
            group=_to_text(row.get("group")),
            currency=_to_text(row.get("currency")),
            counterparty=_to_text(row.get("counterparty")),
            business_segment=_to_text(row.get("business_segment")),
            strategic_segment=_to_text(row.get("strategic_segment")),
            book_value_def=_to_text(row.get("book_value_def")),
            rate_type=_to_text(row.get("rate_type")),
            maturity_bucket=_to_text(row.get("maturity_bucket")),
            remuneration_bucket=_to_text(row.get("remuneration_bucket")),
            maturity_years=_to_float(row.get("maturity_years")),
            amount=_to_float(row.get("amount")),
            rate=_to_float(row.get("rate_display")),
        )
        for row in sliced_records
    ]

    return BalanceContractsResponse(
        session_id=session_id,
        total=total,
        page=effective_page,
        page_size=effective_page_size,
        contracts=contracts,
    )


# ── Excel export ───────────────────────────────────────────────────────────

_EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("Contract ID", "contract_id"),
    ("Sheet", "sheet"),
    ("Category", "categoria_ui"),
    ("Subcategory", "subcategoria_ui"),
    ("Group", "group"),
    ("Currency", "currency"),
    ("Rate Type", "rate_type"),
    ("Segment", "business_segment"),
    ("Book Value Def", "book_value_def"),
    ("Maturity Bucket", "maturity_bucket"),
    ("Remuneration Bucket", "remuneration_bucket"),
    ("Amount", "amount"),
    ("Rate (%)", "rate_display"),
    ("Maturity (Years)", "maturity_years"),
]


def _sanitize_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no special chars."""
    for ch in "[]:*?/\\":
        name = name.replace(ch, "_")
    return name[:31] or "Sheet"


def _write_export_sheet(
    wb, sheet_name: str, df: pd.DataFrame, export_cols: list[tuple[str, str]],
) -> None:
    """Write a single sheet with headers, data rows, and a summary row."""
    from openpyxl.styles import Font, numbers

    ws = wb.create_sheet(title=_sanitize_sheet_name(sheet_name))

    headers = [h for h, _ in export_cols]
    cols = [c for _, c in export_cols]
    available = [c for c in cols if c in df.columns]
    header_map = {c: h for h, c in export_cols}

    # Write header row
    for ci, col in enumerate(available, start=1):
        cell = ws.cell(row=1, column=ci, value=header_map[col])
        cell.font = Font(bold=True)

    # Write data rows
    for ri, (_, row) in enumerate(df.iterrows(), start=2):
        for ci, col in enumerate(available, start=1):
            val = row.get(col)
            if pd.isna(val):
                continue
            if col == "rate_display":
                val = float(val) * 100  # decimal → percentage
            elif col in ("amount", "maturity_years"):
                val = float(val)
            else:
                val = str(val)
            ws.cell(row=ri, column=ci, value=val)

    # Number formatting for Amount and Rate columns
    for ci, col in enumerate(available, start=1):
        if col == "amount":
            for ri in range(2, len(df) + 2):
                ws.cell(row=ri, column=ci).number_format = '#,##0'
        elif col == "rate_display":
            for ri in range(2, len(df) + 2):
                ws.cell(row=ri, column=ci).number_format = '0.00'
        elif col == "maturity_years":
            for ri in range(2, len(df) + 2):
                ws.cell(row=ri, column=ci).number_format = '0.0'

    # Summary row
    summary_row = len(df) + 2
    ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
    for ci, col in enumerate(available, start=1):
        if col == "amount":
            total_amount = float(df["amount"].sum()) if "amount" in df.columns else 0
            cell = ws.cell(row=summary_row, column=ci, value=total_amount)
            cell.font = Font(bold=True)
            cell.number_format = '#,##0'
    # Position count next to Contract ID header
    ws.cell(row=summary_row, column=2, value=f"{len(df)} positions").font = Font(bold=True)

    # Auto-filter
    if available:
        from openpyxl.utils import get_column_letter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(available))}{len(df) + 1}"


@router.get("/api/sessions/{session_id}/balance/export")
def export_balance(
    session_id: str,
    categoria_ui: str | None = None,
    subcategoria_ui: str | None = None,
    subcategory_id: str | None = None,
    currency: str | None = None,
    rate_type: str | None = None,
    counterparty: str | None = None,
    segment: str | None = None,
    strategic_segment: str | None = None,
    maturity: str | None = None,
    remuneration: str | None = None,
    book_value: str | None = None,
    group_by: str | None = None,
):
    """Export filtered balance positions as an Excel (.xlsx) file."""
    from datetime import date

    from openpyxl import Workbook

    _assert_session_exists(session_id)

    df = _load_or_rebuild_positions_df(session_id)

    # Context filters (category / subcategory)
    context_df = _apply_filters_df(
        df,
        categoria_ui=categoria_ui,
        subcategoria_ui=subcategoria_ui,
        subcategory_id=subcategory_id,
    )

    # Detail filters
    filtered_df = _apply_filters_df(
        context_df,
        currency=currency,
        rate_type=rate_type,
        counterparty=counterparty,
        segment=segment,
        strategic_segment=strategic_segment,
        maturity=maturity,
        remuneration=remuneration,
        book_value=book_value,
    )

    if filtered_df.empty:
        raise HTTPException(status_code=404, detail="No positions match the current filters")

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    group_by_cols = [c.strip() for c in group_by.split(",")] if group_by else None

    if group_by_cols:
        valid_cols = [c for c in group_by_cols if c in filtered_df.columns]
        if valid_cols:
            if len(valid_cols) == 1:
                grp_series = filtered_df[valid_cols[0]].fillna("Ungrouped")
            else:
                grp_series = filtered_df[valid_cols].fillna("—").apply(
                    lambda row: " | ".join(str(v) for v in row), axis=1
                )
            filtered_df = filtered_df.copy()
            filtered_df["_export_group"] = grp_series
            for group_name, group_df in filtered_df.groupby("_export_group", sort=True):
                _write_export_sheet(wb, str(group_name), group_df, _EXPORT_COLUMNS)
        else:
            _write_export_sheet(wb, "Positions", filtered_df, _EXPORT_COLUMNS)
    else:
        _write_export_sheet(wb, "Positions", filtered_df, _EXPORT_COLUMNS)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = date.today().isoformat()
    filename = f"balance_export_{session_id[:8]}_{today}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
